import streamlit as st
import pandas as pd
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    st.error("❌ openpyxl 이 설치되지 않았습니다.  python -m pip install openpyxl")
    st.stop()

# ✅ 고정값(입력 X) — 결과에서만 표시
OJC_EXTRA_M = 2.0
LOSS_DB_PER_M = 0.0003

# =========================
# 설정
# =========================
DRAWINGS_DIR = Path(__file__).parent / "drawings"

# ✅ 장비 인식: (DU 제거) LTEDU, 5GDU, LTEMUX 포함
EQUIP_PATTERN = re.compile(
    r"^(RACK|FDF|MUX|LTEMUX|TIE|LTEDU|5GDU)([\s\-]*\d+)?$",
    re.IGNORECASE
)

# =========================
# OJC 규격 매핑 (사용자 제공 표 기준)
# - 대칭(순서 무관)으로 처리
# =========================
def _k(a: str, b: str) -> tuple:
    return tuple(sorted([a, b]))

OJC_MAP = {
    _k("5GDU", "MUX"):     "LC(PC)-LC(PC) 2CORE",
    _k("5GDU", "FDF"):     "LC(PC)-SC(PC) 2CORE",

    _k("LTEDU", "MUX"):    "LC(PC)-SC(PC) 1CORE",
    _k("LTEDU", "LTEMUX"): "SC(PC)-LC(PC) 1CORE",
    _k("LTEDU", "FDF"):    "SC(PC)-LC(PC) 1CORE",
    _k("LTEDU", "TIE"):    "SC(PC)-LC(PC) 1CORE",

    _k("MUX", "LTEMUX"):   "SC(PC)-SC(PC) 1CORE",
    _k("MUX", "FDF"):      "SC(PC)-SC(PC) 1CORE",
    _k("MUX", "TIE"):      "SC(PC)-SC(PC) 1CORE",

    _k("LTEMUX", "FDF"):   "SC(PC)-SC(PC) 1CORE",
    _k("LTEMUX", "TIE"):   "SC(PC)-SC(PC) 1CORE",

    _k("FDF", "TIE"):      "SC(PC)-SC(PC) 1CORE",

    # 표에 FDF-FDF도 값이 있었음(원하면 유지)
    _k("FDF", "FDF"):      "SC(PC)-SC(PC) 1CORE",
}

def get_ojc_spec(kind1: str, kind2: str) -> str:
    """장비종류 2개로 OJC 규격 문자열 반환. 없으면 '-'"""
    return OJC_MAP.get(_k(kind1, kind2), "-")

# =========================
# 세션(누적 결과)
# =========================
if "saved_results" not in st.session_state:
    st.session_state.saved_results = []  # list of dict

# =========================
# UI
# =========================
st.set_page_config(page_title="집중국사 OJC Length 자동계산기", layout="wide")
st.title("📐 집중국사 OJC Length 자동계산기")
st.caption("엑셀 도면 기반 · 표 중심 · 직각거리 기준 · OJC 길이 및 손실 자동 계산")

# =========================
# 입력 파라미터 (셀 크기만)
# =========================
st.subheader("⚙️ 계산 파라미터")

unit = st.radio("1칸(셀) 길이 입력 단위", ["cm", "m"], horizontal=True)

c1, c2 = st.columns(2)
with c1:
    cell_w = st.number_input("1칸 가로 길이", min_value=0.0, value=50.0 if unit == "cm" else 0.5)
with c2:
    cell_h = st.number_input("1칸 세로 길이", min_value=0.0, value=50.0 if unit == "cm" else 0.5)

def to_m(v: float) -> float:
    return v / 100.0 if unit == "cm" else v

cell_w_m = to_m(cell_w)
cell_h_m = to_m(cell_h)

if cell_w_m <= 0 or cell_h_m <= 0:
    st.error("1칸 가로/세로 길이는 0보다 커야 합니다.")
    st.stop()

# 고정값(계산용)
ojc_extra_m = OJC_EXTRA_M
loss_db_per_m = LOSS_DB_PER_M

# =========================
# 도면 선택
# =========================
st.subheader("📂 도면 선택")
mode = st.radio("도면 가져오기 방식", ["폴더에서 선택(drawings)", "업로드"], horizontal=True)

file_obj = None
selected_name = None

if mode == "폴더에서 선택(drawings)":
    DRAWINGS_DIR.mkdir(exist_ok=True)
    files = sorted(DRAWINGS_DIR.glob("*.xlsx"))
    if not files:
        st.warning("drawings 폴더에 도면(.xlsx)이 없습니다.")
        st.stop()
    path = st.selectbox("도면 파일 선택", files, format_func=lambda p: p.name)
    file_obj, selected_name = path, path.name
else:
    up = st.file_uploader("엑셀 도면 업로드(.xlsx)", type=["xlsx"])
    if not up:
        st.stop()
    file_obj, selected_name = up, up.name

# =========================
# 엑셀 로딩
# =========================
try:
    wb = load_workbook(file_obj, data_only=True)
except Exception as e:
    st.error(f"엑셀 파일을 열 수 없습니다: {e}")
    st.stop()

sheet = st.selectbox("시트 선택", wb.sheetnames)
ws = wb[sheet]

# =========================
# 장비 추출
# =========================
items = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None:
            continue

        raw = str(cell.value).strip().upper()
        if not EQUIP_PATTERN.match(raw):
            continue

        # 공백/하이픈 제거: "5GDU-01" / "LTEMUX 01" -> "5GDU01" / "LTEMUX01"
        norm = re.sub(r"[\s\-]+", "", raw)

        kind_m = re.match(r"^(RACK|FDF|MUX|LTEMUX|TIE|LTEDU|5GDU)", norm)
        if not kind_m:
            continue
        kind = kind_m.group(1)

        num = norm[len(kind):]
        name = f"{kind}{num}" if num else kind

        items.append({
            "장비종류": kind,
            "장비명": name,
            "엑셀행": cell.row,
            "엑셀열": cell.column
        })

if not items:
    st.error("❗ 도면에서 장비(RACK/FDF/MUX/LTEMUX/TIE/LTEDU/5GDU)를 찾지 못했습니다.")
    st.stop()

df = pd.DataFrame(items).drop_duplicates("장비명").copy()

# =========================
# 좌표 계산 (기준점: 좌상단 장비)
# =========================
min_row, min_col = df["엑셀행"].min(), df["엑셀열"].min()
df["오른쪽으로_몇칸"] = df["엑셀열"] - min_col
df["아래로_몇칸"] = df["엑셀행"] - min_row

# =========================
# 장비 위치 목록 (슬림 + 검색 + 정렬 + 가운데 정렬)
# =========================
st.subheader("📦 장비 위치 목록")

search = st.text_input("장비명 검색 (예: 5GDU, LTEDU01, LTEMUX01, FDF01, RACK)", "")
df_view = df.copy()
if search.strip():
    df_view = df_view[df_view["장비명"].str.contains(search.strip().upper())]
df_view = df_view.sort_values(by=["장비종류", "장비명"])

table_df = df_view[["장비종류", "장비명", "오른쪽으로_몇칸", "아래로_몇칸"]].copy()

styler = (
    table_df.style
    .set_properties(**{"text-align": "center"})
    .set_table_styles([
        {"selector": "th", "props": [("text-align", "center")]},
        {"selector": "th, td", "props": [("padding", "2px 6px")]},
    ])
)

st.dataframe(
    styler,
    use_container_width=True,
    height=220,
    hide_index=True
)

# =========================
# 공통 계산 함수
# =========================
def calc_pair(a_row: pd.Series, b_row: pd.Series) -> dict:
    dx = abs(int(a_row["오른쪽으로_몇칸"]) - int(b_row["오른쪽으로_몇칸"]))
    dy = abs(int(a_row["아래로_몇칸"]) - int(b_row["아래로_몇칸"]))

    right_angle_m = dx * cell_w_m + dy * cell_h_m
    ojc_m = right_angle_m + ojc_extra_m
    loss_db = ojc_m * loss_db_per_m

    ojc_spec = get_ojc_spec(str(a_row["장비종류"]), str(b_row["장비종류"]))

    return {
        "OJC규격": ojc_spec,                 # ✅ 추가(표 기반)
        "가로차이_칸": dx,
        "세로차이_칸": dy,
        "직각거리(m)": right_angle_m,
        "여장(m)": ojc_extra_m,
        "OJC_길이(m)": ojc_m,
        "손실(dB)": loss_db
    }

# =========================
# 단건 계산 + 누적 저장
# =========================
st.subheader("🎯 단건 계산 (장비 1 ↔ 장비 2)")

names = sorted(df["장비명"].tolist())
if len(names) < 2:
    st.warning("장비가 2개 이상 있어야 계산할 수 있습니다.")
    st.stop()

a_name = st.selectbox("장비 1 선택", names, index=0)
b_name = st.selectbox("장비 2 선택", names, index=1)

if a_name == b_name:
    st.info("서로 다른 장비를 선택하세요.")
    st.stop()

a = df[df["장비명"] == a_name].iloc[0]
b = df[df["장비명"] == b_name].iloc[0]

pair = calc_pair(a, b)

# ✅ OJC 규격 표시(결과)
st.markdown(f"**OJC 규격:** `{pair['OJC규격']}`")
if pair["OJC규격"] == "-":
    st.warning("이 장비 조합은 표 기준 OJC 규격이 없습니다(또는 정의되지 않았습니다).")

c1, c2, c3, c4, c5 = st.columns(5)
c1.metric("가로 차이 (칸)", pair["가로차이_칸"])
c2.metric("세로 차이 (칸)", pair["세로차이_칸"])
c3.metric("직각거리 (m)", f"{pair['직각거리(m)']:.2f}")
c4.metric("OJC 길이 (m)", f"{pair['OJC_길이(m)']:.2f}")
c5.metric("손실 (dB)", f"{pair['손실(dB)']:.6f}")

st.info(f"표준값 적용: OJC 여장 = {ojc_extra_m} m / 손실계수 = {loss_db_per_m} dB/m")
st.caption("손실(dB) = OJC 길이(m) × 손실 계수(dB/m)")

btn1, btn2, btn3 = st.columns(3)
with btn1:
    if st.button("➕ 이 결과를 누적 저장", use_container_width=True):
        st.session_state.saved_results.append({
            "도면파일": selected_name,
            "시트": sheet,
            "장비1": a_name,
            "장비1_종류": str(a["장비종류"]),
            "장비2": b_name,
            "장비2_종류": str(b["장비종류"]),
            "1칸_가로(m)": cell_w_m,
            "1칸_세로(m)": cell_h_m,
            "손실계수(dB/m)": loss_db_per_m,
            **{k: (round(v, 6) if isinstance(v, float) else v) for k, v in pair.items()}
        })
        st.success("누적 저장 완료!")

with btn2:
    if st.button("🧹 누적 결과 지우기", use_container_width=True):
        st.session_state.saved_results = []
        st.success("누적 결과를 비웠습니다.")

with btn3:
    single_df = pd.DataFrame([{
        "도면파일": selected_name,
        "시트": sheet,
        "장비1": a_name,
        "장비1_종류": str(a["장비종류"]),
        "장비2": b_name,
        "장비2_종류": str(b["장비종류"]),
        "1칸_가로(m)": cell_w_m,
        "1칸_세로(m)": cell_h_m,
        "손실계수(dB/m)": loss_db_per_m,
        **{k: (round(v, 6) if isinstance(v, float) else v) for k, v in pair.items()}
    }])
    st.download_button(
        "⬇️ 단건 CSV 다운로드",
        data=single_df.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"ojc_length_{a_name}_{b_name}.csv",
        mime="text/csv",
        use_container_width=True
    )

# =========================
# 장비1 기준 전체 자동 계산표
# =========================
st.subheader("📋 장비 1 기준 전체 자동 계산표")

base_name = st.selectbox("기준 장비(장비 1) 선택", names, key="base_equip")
base = df[df["장비명"] == base_name].iloc[0]

rows = []
for _, r in df.iterrows():
    if r["장비명"] == base_name:
        continue
    out = calc_pair(base, r)
    rows.append({
        "상대장비_종류": r["장비종류"],
        "상대장비": r["장비명"],
        "OJC규격": out["OJC규격"],           # ✅ 추가
        "가로차이(칸)": out["가로차이_칸"],
        "세로차이(칸)": out["세로차이_칸"],
        "직각거리(m)": out["직각거리(m)"],
        "여장(m)": out["여장(m)"],
        "OJC_길이(m)": out["OJC_길이(m)"],
        "손실(dB)": out["손실(dB)"],
        "손실계수(dB/m)": loss_db_per_m,
    })

all_df = pd.DataFrame(rows)
if not all_df.empty:
    all_df = all_df.sort_values(by=["OJC_길이(m)", "상대장비_종류", "상대장비"]).reset_index(drop=True)
    show_df = all_df.copy()
    show_df["직각거리(m)"] = show_df["직각거리(m)"].map(lambda x: round(x, 3))
    show_df["OJC_길이(m)"] = show_df["OJC_길이(m)"].map(lambda x: round(x, 3))
    show_df["손실(dB)"] = show_df["손실(dB)"].map(lambda x: round(x, 6))

    st.dataframe(show_df, use_container_width=True, height=320, hide_index=True)

    st.download_button(
        "⬇️ 기준 장비 전체 결과 CSV 다운로드",
        data=show_df.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"ojc_all_from_{base_name}.csv",
        mime="text/csv"
    )
else:
    st.info("계산할 상대 장비가 없습니다.")

# =========================
# 누적 결과 표시 + 일괄 다운로드
# =========================
st.subheader("📚 누적 결과")

if st.session_state.saved_results:
    saved_df = pd.DataFrame(st.session_state.saved_results)

    # ✅ OJC규격 컬럼을 앞쪽으로 보기 좋게 이동
    front_cols = [c for c in ["OJC규격", "장비1", "장비2"] if c in saved_df.columns]
    rest_cols = [c for c in saved_df.columns if c not in front_cols]
    saved_df = saved_df[front_cols + rest_cols]

    st.dataframe(saved_df, use_container_width=True, height=260, hide_index=True)

    st.download_button(
        "⬇️ 누적 결과 전체 CSV 다운로드",
        data=saved_df.to_csv(index=False).encode("utf-8-sig"),
        file_name="ojc_saved_results.csv",
        mime="text/csv"
    )
else:
    st.info("아직 누적 저장된 결과가 없습니다. 위에서 '➕ 이 결과를 누적 저장'을 눌러주세요.")
